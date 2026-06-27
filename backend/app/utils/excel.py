from io import BytesIO
from fastapi import UploadFile, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook


def parse_excel(file: UploadFile) -> list[dict]:
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be .xlsx or .xls")
    try:
        wb = load_workbook(file.file, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            raise HTTPException(status_code=400, detail="Excel must have header + at least 1 row")
        headers = [str(h).strip() if h else "" for h in rows[0]]
        result = []
        for row in rows[1:]:
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            d = {}
            for i, h in enumerate(headers):
                val = row[i] if i < len(row) else None
                d[h] = str(val).strip() if val is not None else None
            result.append(d)
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel: {str(e)}")


def generate_excel(headers: list[str], rows: list[list]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def excel_response(headers: list[str], rows: list[list], filename: str):
    buf = generate_excel(headers, rows)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
